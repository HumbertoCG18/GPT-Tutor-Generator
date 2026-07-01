#!/usr/bin/env python3
"""Crosswalk gold xlsx (rotulado) -> ground_truth_<curso>.csv. GENERICO por --curso.

Generaliza build_ground_truth_IA.py. Dois MODOS, auto-detectados pelo sheet de
gabarito do xlsx (build_gold_xlsx.py emite um ou outro conforme exista csv de
subtopicos):

  MODO SUBTOPICO  (sheet "Gabarito Subtopicos"):  humano rotula por SUBTOPICO
    (semantico, datado); a maquina deriva o bloco via SARC sob borda [inicio,fim).
    Indirecao robusta: detecta straddle (subtopico atravessa fronteira) e sobrevive
    a renumeracao de bloco. Caso IA.

  MODO BLOCO  (sheet "Gabarito dos Blocos"):  humano escolhe o bloco-NN direto no
    dropdown; true_block_id = bloco_correto, sem crosswalk de data. Caso dos cursos
    sem csv de subtopicos (ES2/MF/SO/TCC) ate que se autore o csv.

Regras TRAVADAS (herdadas da sessao 2026-06-25, validas para o modo subtopico):
  - Borda [inicio, fim): esquerda-inclusiva, direita-exclusiva. data == inicio_do_bloco
    cai NESSE bloco, nao no anterior. Teste-unidade dedicado, LOAD-BEARING.
  - predicted = TEMPORAL (temporal_block_id vence; fallback manual>computed). Espelha
    eval_ground_truth.load_predictions.
  - data_real = oraculo SARC independente (Gabarito Subtopicos). NAO card/posting.
  - escopo e dedup (pairs) sao CONFIG POR-CURSO (decisao humana; ver COURSE_CONFIG).
  - denominador DERIVADO explicito. NAO commita; HALT pra revisao humana do crosswalk.

Emite (uncommitted): id, true_block_id, computed_block_id, temporal_block_id, pair_key,
                     provenance, scope, data_real, scorable, discriminante.

Uso:  python scripts/build_ground_truth.py --curso IA
      python scripts/build_ground_truth.py --curso SO --out C:/tmp/gt_SO.csv
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# CURSO_REPOS / GITHUB_BASE vivem no gerador de scaffold — fonte unica, sem drift.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_gold_xlsx import CURSO_REPOS, GITHUB_BASE  # noqa: E402

# stdout UTF-8: console cp1252 (Windows) crasha em char fora do Latin-1.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ============================ CONFIG POR-CURSO ===============================
# So o modo subtopico usa clean/eco/year. `pairs` (dedup de version-pairs) vale
# nos dois modos. Curso novo em modo-bloco: comece com pairs={} e ajuste quando
# os pares dup forem conhecidos (NUNCA derivado de bloco computado).
COURSE_CONFIG = {
    "IA": {
        # subtopicos letivos pontuaveis: 1-11, 16-20. 12-15 = ECO (fonte "Card SNN")
        # excluido e carimbado, nao entra no denominador.
        "clean_subtopics": set(range(1, 12)) | set(range(16, 21)),
        "eco_subtopics": {12, 13, 14, 15},
        # 3 pares dedup md5 (manifest ids). pair_key = canonico.
        "pairs": {
            "minimax-teoria": "minimax", "minimax": "minimax",
            "lista1": "lista-de-exercicios-i", "lista-de-exercicios-i": "lista-de-exercicios-i",
            "prova-1-2024-02": "prova-1-2024-02", "prova-1-202402": "prova-1-2024-02",
        },
        "year": 2026,
    },
    # --- modo subtopico (sarc_subtopics_<curso>.csv existe; encaixe limpo) ------
    # clean/eco DERIVADOS do csv: letivo=yes e fonte sem "Card" -> clean; "Card" -> eco.
    # Regra reproduz o IA byte-a-byte. ACOPLADO a ordem das linhas do csv: se
    # reordenar/fundir, re-derive (scripts/../draft_subtopics.py imprime o set).
    # pairs placeholder: preencha quando souber os dups (NUNCA de bloco computado).
    "ES2": {"clean_subtopics": set(range(1, 18)) - {5, 9}, "eco_subtopics": set(), "pairs": {}, "year": 2026},
    # --- modo-bloco (placeholders; preencha pairs quando souber os dups) -------
    # MF: subtopico testado mas 10/18 semanas STRADDLE (blocos do MF nao alinham
    # com a semana do roteiro) -> forcado a bloco via FORCE_BLOCK em build_gold_xlsx.
    # clean/eco inertes em modo-bloco; sarc_subtopics_MF.csv fica so como registro.
    # MF: 1 par md5-provado via raw/ (sweep 2026-07-01); gemeos ambos bloco-10 no gold
    # (consistentes, ambos PASS) — sem flip de veredito, so dedup de denominador.
    "MF":  {"clean_subtopics": set(), "eco_subtopics": set(), "pairs": {
        "logicadehoare1-exercicios-respostas": "logicadehoare-exercicios-respostas",
        "logicadehoare-exercicios-respostas": "logicadehoare-exercicios-respostas",
    }, "year": 2026},
    # SO: dups md5-provados via raw/ (sweep 2026-07-01; 42 entries = 38 distintos).
    # Fonte 100% Moodle (sem stash antigo) — 4 pares intra-Moodle, rotulos dos gemeos
    # consistentes no gold (validado). Canonico = id mais descritivo.
    "SO":  {"clean_subtopics": set(), "eco_subtopics": set(), "pairs": {
        "lista1-gab": "lista-exercicios-p1-gabarito",
        "lista-exercicios-p1-gabarito": "lista-exercicios-p1-gabarito",
        "lista2": "lista-exercicios-p2",
        "lista-exercicios-p2": "lista-exercicios-p2",
        "programa": "plano-de-ensino",
        "plano-de-ensino": "plano-de-ensino",
        "14-04-troca-de-mensagens": "1404-troca-de-mensagens",
        "1404-troca-de-mensagens": "1404-troca-de-mensagens",
    }, "year": 2026},
    # TCC: dups md5-provados via raw/ do repo (sweep 2026-07-01; 42 entries = 27 distintos).
    # Causa: stash antigo Downloads/TCC (24 entries, sources ja SUMIDOS do disco) acumulado
    # com o stash Moodle (18) sem poda de migracao — mesmo mecanismo do IA (2026-06-23).
    # Canonico = id do lado MOODLE (vivo) quando existe, pra sobreviver a poda futura.
    # ATENCAO: aula-06 tem COLISAO DE ID no manifest (2 entries com o mesmo id) + 3a copia
    # com id proprio; as 3 sao byte-identicas e mapeiam ao mesmo canonico.
    "TCC": {"clean_subtopics": set(), "eco_subtopics": set(), "pairs": {
        # --- 11 pares cross-stash (OLD ≡ Moodle, byte-identico) ---
        "enunciado-t2": "trabalho-t2-enunciado",
        "trabalho-t2-enunciado": "trabalho-t2-enunciado",
        "aula-13-teorema-de-rice": "aula-13-teorema-de-rice-pdf",
        "aula-13-teorema-de-rice-pdf": "aula-13-teorema-de-rice-pdf",
        "aula-11-o-problema-da-parada-halting-problem-halteproblem":
            "aula-11-o-problema-da-parada-halting-problem-halteproblem-pdf",
        "aula-11-o-problema-da-parada-halting-problem-halteproblem-pdf":
            "aula-11-o-problema-da-parada-halting-problem-halteproblem-pdf",
        "aula-08-maquinas-de-turing-como-processadores-de-funcoes":
            "aula-08-maquinas-de-turing-como-processadoras-de-funcoes",
        "aula-08-maquinas-de-turing-como-processadoras-de-funcoes":
            "aula-08-maquinas-de-turing-como-processadoras-de-funcoes",
        "aula-12-entscheidungsproblem": "aula-12-entscheidungsproblem-pdf",
        "aula-12-entscheidungsproblem-pdf": "aula-12-entscheidungsproblem-pdf",
        "aula-04-funcoes-computaveis-funcoes-recursivas-parciais":
            "aula-04-funcoes-computaveis-funcoes-recursivas-parciais-pdf",
        "aula-04-funcoes-computaveis-funcoes-recursivas-parciais-pdf":
            "aula-04-funcoes-computaveis-funcoes-recursivas-parciais-pdf",
        "aula-17-np-completude": "aula-17-np-completude-pdf",
        "aula-17-np-completude-pdf": "aula-17-np-completude-pdf",
        "enunciado-t1": "t1-enunciado",
        "t1-enunciado": "t1-enunciado",
        "aula-16-classes-de-problemas-e-complexidade":
            "aula-16-classes-de-problemas-e-complexidade-pdf",
        "aula-16-classes-de-problemas-e-complexidade-pdf":
            "aula-16-classes-de-problemas-e-complexidade-pdf",
        "aula07-maquinas-de-turing-e-linguagens-recursivamente-enumeraveis":
            "aula-07-maquinas-de-turing-e-linguagens-recursivamente-enumeraveis",
        "aula-07-maquinas-de-turing-e-linguagens-recursivamente-enumeraveis":
            "aula-07-maquinas-de-turing-e-linguagens-recursivamente-enumeraveis",
        # ids carregam U+0131 (dotless-i, NFD macOS) — copiar byte-a-byte, nao "corrigir".
        "aula-10-linguagens-reconhecıveis-e-linguagens-decidıveis":
            "aula-10-linguagens-reconhecıveis-e-linguagens-decidıveis-pdf",
        "aula-10-linguagens-reconhecıveis-e-linguagens-decidıveis-pdf":
            "aula-10-linguagens-reconhecıveis-e-linguagens-decidıveis-pdf",
        # --- triplo intra-OLD (aula-06 x3; sem copia Moodle) ---
        "aula-06-revisao-alfabeto-cadeia-linguagem-hierarquia-de-chomsky-lemas-e-propriedades-de-automatos":
            "aula-06-revisao-alfabeto-cadeia-linguagem-hierarquia-de-chomsky-lemas-e-propriedades-de-automatos",
        "revisao-alfabeto-cadeia-linguagem-hierarquia-de-chomsky-lemas-e-propriedades-de-automatos":
            "aula-06-revisao-alfabeto-cadeia-linguagem-hierarquia-de-chomsky-lemas-e-propriedades-de-automatos",
        # --- 2 pares intra-Moodle (mesmo PDF postado 2x com nomes diferentes) ---
        "3dm-caetano-gabriel-e-gustavo": "3d-matching",
        "3d-matching": "3d-matching",
        "programacao-inteira-01-20260617-154423-0000": "integer-programming-0001",
        "integer-programming-0001": "integer-programming-0001",
    }, "year": 2026},
}

OUT_COLS = ["id", "material", "true_block_id", "computed_block_id", "temporal_block_id",
            "pair_key", "provenance", "scope", "data_real", "scorable", "discriminante"]
BAR = "=" * 72


def _nalnum(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# ============================ infra compartilhada ============================
class Course:
    """Fontes READ-ONLY do repo + helpers de bloco/temporal/join (os dois modos usam)."""

    def __init__(self, repo: Path, year: int):
        self.repo = repo
        self.year = year
        tl = self._load("course/.timeline_index.json")
        blocks = tl if isinstance(tl, list) else tl.get("blocks", [])
        self.blocks = sorted(blocks, key=lambda b: b.get("period_start", ""))
        self.starts = [(b.get("id"), b.get("period_start")) for b in self.blocks]
        self.valid_blocks = {str(b.get("id")) for b in self.blocks}
        bi = self._load("course/.block_identity.json")
        self.u2d = {str(b.get("uuid")): str(b.get("display_id_last")) for b in bi}
        self._build_join_index()

    def _load(self, rel: str):
        return json.loads((self.repo / rel).read_text(encoding="utf-8"))

    def iso(self, dd_mm: str) -> str:
        d, m = dd_mm.split("/")
        return f"{self.year}-{int(m):02d}-{int(d):02d}"

    def block_for_date(self, d_iso: str):
        """Borda [inicio, proximo inicio). data==inicio cai NESSE bloco."""
        cur = None
        for bid, st in self.starts:
            if st <= d_iso:
                cur = bid
            else:
                break
        return cur

    def border_unit_test(self) -> list:
        """LOAD-BEARING: toda data==inicio cai no proprio bloco, nao no anterior."""
        return [(bid, st, self.block_for_date(st))
                for bid, st in self.starts if self.block_for_date(st) != bid]

    def disp(self, uuid_or_id) -> str:
        s = str(uuid_or_id or "").strip()
        return self.u2d.get(s, s)

    def predicted_temporal(self, e: dict) -> str:
        t = str(e.get("temporal_block_id") or "").strip()
        if t:
            return self.disp(t)
        man = str(e.get("manual_timeline_block_id") or "").strip()
        if man:
            return self.disp(man)
        return self.disp(e.get("computed_block_id"))

    def _build_join_index(self):
        m = self._load("manifest.json")
        self.by_base, self.by_title, self.by_nbase, self.by_ntitle = {}, {}, {}, {}
        for e in m["entries"]:
            sp = str(e.get("source_path") or "")
            self.by_base[os.path.basename(sp).lower()] = e
            self.by_title[str(e.get("title") or "").lower()] = e
            self.by_nbase[_nalnum(os.path.basename(sp))] = e
            self.by_ntitle[_nalnum(e.get("title"))] = e

    def join(self, material):
        """EXATO (basename/title) -> NORMALIZADO-alnum. Sem match = nao na manifest
        (podado/renomeado) -> unjoined honesto, sem force-match de fuzzy."""
        k = str(material or "").strip().lower()
        nk = _nalnum(material)
        return (self.by_base.get(k) or self.by_title.get(k)
                or self.by_nbase.get(nk) or self.by_ntitle.get(nk))


# ============================ MODO SUBTOPICO (IA) ============================
def run_subtopic(co: Course, wb, cfg: dict) -> tuple:
    clean, eco = cfg["clean_subtopics"], cfg["eco_subtopics"]
    pairs = cfg["pairs"]

    # Gabarito Subtopicos: subtopico# -> datas, fonte
    gab = list(wb["Gabarito Subtopicos"].iter_rows(values_only=True))[1:]
    sub_dates, sub_fonte = {}, {}
    for r in gab:
        if r[0] is None:
            continue
        n = int(r[0])
        sub_dates[n] = re.findall(r"\d{1,2}/\d{1,2}", str(r[2] or ""))
        sub_fonte[n] = str(r[4] or "")

    # CROSSWALK subtopico -> {blocos} via datas sob [inicio,fim)
    crosswalk = {}
    for n, dates in sorted(sub_dates.items()):
        bset = [co.block_for_date(co.iso(d)) for d in dates]
        crosswalk[n] = {"dates": dates, "blocks": bset,
                        "single": len(set(bset)) == 1, "fonte": sub_fonte[n]}

    rot = list(wb["Rotulagem"].iter_rows(values_only=True))[1:]
    out_rows, unjoined, straddle_flag = [], [], []
    for r in rot:
        if not any(c not in (None, "") for c in r):
            continue
        material = r[1]
        if material is None:
            continue
        bloco_correto = str(r[4] or ""); obs = str(r[6] or "")
        mnum = re.match(r"\s*(\d+)", bloco_correto)
        if mnum:
            sub = int(mnum.group(1))
        else:
            ms = re.search(r"subtopico\s+(\d+)", obs)
            sub = int(ms.group(1)) if ms else None
        scope = "clean" if (sub in clean) else ("eco-excluido" if sub in eco else "sem-subtopico")
        fonte = sub_fonte.get(sub, "")

        md = re.search(r"->\s*(\d{1,2}/\d{1,2})", obs)
        prov = "?"
        if md:
            data_real = md.group(1); true_block = co.block_for_date(co.iso(data_real)); prov = f"obs:{fonte}"
        elif sub in crosswalk:
            cw = crosswalk[sub]
            if cw["single"]:
                data_real = cw["dates"][0] if cw["dates"] else ""; true_block = cw["blocks"][0]; prov = f"subt-single:{fonte}"
            else:
                data_real = ""; true_block = ""; prov = f"subt-STRADDLE:{fonte}"
        else:
            data_real = ""; true_block = ""

        e = co.join(material)
        eid = str(e.get("id")) if e else ""
        comp = co.disp(e.get("computed_block_id")) if e else ""
        temp = co.predicted_temporal(e) if e else ""
        pair = pairs.get(eid, "")
        if not e:
            unjoined.append(material)

        is_straddle = prov.startswith("subt-STRADDLE")
        scorable = bool(e) and scope == "clean" and bool(true_block) and not is_straddle
        discrim = scorable and (true_block != temp)
        if is_straddle and scope == "clean":
            straddle_flag.append((material, sub))

        out_rows.append({
            "id": eid, "material": material, "true_block_id": (true_block if scorable else ""),
            "computed_block_id": comp, "temporal_block_id": temp, "pair_key": pair,
            "provenance": prov, "scope": scope, "data_real": data_real,
            "scorable": "yes" if scorable else "no",
            "discriminante": "yes" if discrim else "no",
        })

    # --- CROSSWALK display ---
    print("\n" + BAR); print("CROSSWALK  subtopico -> bloco(s)  (sob [inicio,fim))"); print(BAR)
    for n in sorted(crosswalk):
        cw = crosswalk[n]
        tag = "single " if cw["single"] else "STRADDLE"
        sc = "clean" if n in clean else ("ECO" if n in eco else "-")
        print(f"  subt {n:2} [{sc:5}] {tag} | datas {cw['dates']} -> {cw['blocks']} | fonte: {cw['fonte']}")

    extras = {"straddle_flag": straddle_flag, "eco": eco}
    return out_rows, unjoined, extras


# ============================ MODO BLOCO ====================================
def run_block(co: Course, wb, cfg: dict) -> tuple:
    pairs = cfg["pairs"]
    rot = list(wb["Rotulagem"].iter_rows(values_only=True))[1:]
    out_rows, unjoined = [], []
    for r in rot:
        if not any(c not in (None, "") for c in r):
            continue
        material = r[1]
        if material is None:
            continue
        bloco = str(r[4] or "").strip(); obs = str(r[6] or "")
        e = co.join(material)
        eid = str(e.get("id")) if e else ""
        comp = co.disp(e.get("computed_block_id")) if e else ""
        temp = co.predicted_temporal(e) if e else ""
        pair = pairs.get(eid, "")
        if not e:
            unjoined.append(material)

        # bloco_correto = bloco-NN direto (dropdown), "N/A", "nao sei" ou vazio.
        if bloco in co.valid_blocks:
            true_block = bloco; scope = "clean"; prov = "block-direct"
        elif _nalnum(bloco) == "na":
            true_block = ""; scope = "na"; prov = "block-direct:NA"
        else:                                   # "nao sei" / vazio / valor estranho
            true_block = ""; scope = "sem-rotulo"; prov = "block-direct"

        scorable = bool(e) and scope == "clean" and bool(true_block)
        discrim = scorable and (true_block != temp)

        out_rows.append({
            "id": eid, "material": material, "true_block_id": (true_block if scorable else ""),
            "computed_block_id": comp, "temporal_block_id": temp, "pair_key": pair,
            "provenance": prov, "scope": scope, "data_real": "",
            "scorable": "yes" if scorable else "no",
            "discriminante": "yes" if discrim else "no",
        })
    return out_rows, unjoined, {"straddle_flag": [], "eco": set()}


# ============================ denominador + emit ============================
def report_and_write(curso: str, mode: str, out_rows: list, unjoined: list,
                     extras: dict, out_path: Path) -> None:
    straddle_flag, eco = extras["straddle_flag"], extras["eco"]
    scor = [x for x in out_rows if x["scorable"] == "yes"]
    disc = [x for x in scor if x["discriminante"] == "yes"]

    print("\n" + BAR); print(f"DENOMINADOR DERIVADO (explicito) — curso={curso} modo={mode}"); print(BAR)
    print(f"  linhas Rotulagem            : {len(out_rows)}")
    print(f"  joined ao manifest          : {sum(1 for x in out_rows if x['id'])}/{len(out_rows)}  (unjoined: {len(unjoined)})")
    print(f"  escopo clean                : {sum(1 for x in out_rows if x['scope']=='clean')}")
    if mode == "subtopico":
        print(f"  eco-excluido ({sorted(eco)}) : {sum(1 for x in out_rows if x['scope']=='eco-excluido')}")
        print(f"  straddle-flag (clean s/data): {len(straddle_flag)}  <- precisam data_real por-material [USER]")
    else:
        print(f"  N/A (fora da timeline)      : {sum(1 for x in out_rows if x['scope']=='na')}")
        print(f"  sem-rotulo (pendente USER)  : {sum(1 for x in out_rows if x['scope']=='sem-rotulo')}")
    print(f"  >>> SCORABLE (denominador)  : {len(scor)}")
    print(f"  >>> DISCRIMINANTES          : {len(disc)}")

    print("\n  DISCRIMINANTES nomeados (true != temporal):")
    for x in disc:
        print(f"    FAIL {x['material'][:46]:46} data_real {x['data_real']} -> true {x['true_block_id']} | temporal {x['temporal_block_id']}")

    if straddle_flag:
        print("\n  STRADDLE-FLAG (clean, subtopico atravessa fronteira, sem data_real por-material):")
        for mat, sub in straddle_flag:
            print(f"    subt {sub} | {mat[:54]}")

    if unjoined:
        print("\n  UNJOINED (material sem match no manifest -- sem gap silencioso):")
        for u in unjoined:
            print(f"    {u}")

    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS); w.writeheader()
        for x in out_rows:
            w.writerow(x)

    print("\n" + BAR)
    print(f"HALT pos-crosswalk. CSV escrito (NAO commitado): {out_path}")
    print("Revisa o crosswalk + denominador acima ANTES de rodar o eval. Nada commitado.")
    print(BAR)


def main(argv: list) -> int:
    ap = argparse.ArgumentParser(description="Crosswalk gold xlsx -> ground_truth_<curso>.csv.")
    ap.add_argument("--curso", required=True, help="IA / ES2 / MF / SO / TCC")
    ap.add_argument("--repo", default="", help="override do caminho do repo-tutor")
    ap.add_argument("--xlsx", default="", help="override do xlsx rotulado")
    ap.add_argument("--out", default="", help="override do csv de saida")
    args = ap.parse_args(argv)

    curso = args.curso.strip().upper()
    cfg = COURSE_CONFIG.get(curso)
    if cfg is None:
        print(f"[erro] curso {curso!r} sem config; conhecidos: {list(COURSE_CONFIG)}")
        return 2

    if args.repo:
        repo = Path(args.repo)
    elif curso in CURSO_REPOS:
        repo = GITHUB_BASE / CURSO_REPOS[curso]
    else:
        print(f"[erro] curso {curso!r} sem repo conhecido; use --repo.")
        return 2
    if not (repo / "manifest.json").exists():
        print(f"[erro] manifest.json nao encontrado em {repo}")
        return 2

    templates = Path(__file__).resolve().parents[1] / "docs" / "reports" / "gold_templates"
    xlsx = Path(args.xlsx) if args.xlsx else (templates / f"gold_{curso}_rotular.xlsx")
    if not xlsx.exists():
        print(f"[erro] xlsx rotulado nao encontrado: {xlsx}")
        return 2
    out_path = Path(args.out) if args.out else (
        Path(__file__).resolve().parents[1] / "docs" / "reports" / f"ground_truth_{curso}.csv")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    co = Course(repo, cfg["year"])

    # --- TESTE-UNIDADE da borda (LOAD-BEARING) ---
    print(BAR); print(f"TESTE-UNIDADE BORDA [inicio, fim)  — curso={curso}"); print(BAR)
    fails = co.border_unit_test()
    if fails:
        print("FALHOU -- a borda nao e [inicio,fim). HALT.")
        for bid, st, got in fails:
            print(f"  data {st} (inicio de {bid}) caiu em {got}")
        return 1
    print(f"OK: {len(co.starts)} blocos, toda data==inicio cai no proprio bloco.")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    sheets = set(wb.sheetnames)
    if "Gabarito Subtopicos" in sheets:
        mode = "subtopico"
        out_rows, unjoined, extras = run_subtopic(co, wb, cfg)
    elif "Gabarito dos Blocos" in sheets:
        mode = "bloco"
        out_rows, unjoined, extras = run_block(co, wb, cfg)
    else:
        print(f"[erro] xlsx sem sheet de gabarito reconhecido. sheets={sorted(sheets)}")
        return 2

    report_and_write(curso, mode, out_rows, unjoined, extras, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
