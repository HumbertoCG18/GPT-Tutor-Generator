#!/usr/bin/env python3
"""Conversor gold xlsx -> ground_truth_IA.csv. HALT pos-crosswalk (read-before-commit).

Regras TRAVADAS (sessao 2026-06-25):
  - Borda [inicio, fim): esquerda-inclusiva, direita-exclusiva. data == inicio_do_bloco
    cai NESSE bloco, nao no anterior. Decisao GLOBAL/principled (1 bloco/data = funcao).
    -> teste-unidade dedicado, LOAD-BEARING (fiador dos 2 FAILs k-NN de fronteira).
  - predicted = TEMPORAL (resolve_temporal_block: temporal_block_id vence; fallback
    manual>computed). Espelha eval_ground_truth.load_predictions:81 / file_map:633-636.
  - data_real = oraculo SARC independente (Gabarito Subtopicos). NAO card/posting.
  - escopo: clean = subtopicos 1-11,16-20; ECO 12-15 (fonte "Card SNN") EXCLUIDO, carimbado.
  - 3 pares dedup colapsam por pair_key (regra-c; nao-cascateamento).
  - denominador DERIVADO explicito. NAO commita; HALT pra revisao humana do crosswalk.

Emite (uncommitted): id, true_block_id, computed_block_id, temporal_block_id, pair_key,
                     provenance, scope, scorable, discriminante.

Uso:  python scripts/build_ground_truth_IA.py
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys

import openpyxl

# stdout UTF-8: console cp1252 (Windows) crasha em char fora do Latin-1 (ex.: titulo TCC).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

REPO = r"C:/Users/Humberto/Documents/GitHub/Inteligencia-Artifical-Tutor"
XLSX = r"C:/Users/Humberto/Documents/GitHub/GPT-Tutor-Generator/docs/reports/gold_templates/gold_IA_rotular.xlsx"
OUT = r"C:/Users/Humberto/Documents/GitHub/GPT-Tutor-Generator/docs/reports/ground_truth_IA.csv"

CLEAN_SUBTOPICS = set(range(1, 12)) | set(range(16, 21))   # 1-11, 16-20
# 3 pares dedup md5 (manifest ids). pair_key = canonico.
PAIRS = {
    "minimax-teoria": "minimax", "minimax": "minimax",
    "lista1": "lista-de-exercicios-i", "lista-de-exercicios-i": "lista-de-exercicios-i",
    "prova-1-2024-02": "prova-1-2024-02", "prova-1-202402": "prova-1-2024-02",
}


def load(p):
    return json.load(open(os.path.join(REPO, p), encoding="utf-8"))


def iso(dd_mm):
    d, m = dd_mm.split("/")
    return f"2026-{int(m):02d}-{int(d):02d}"


def main():
    # --- blocos + spans [inicio, proximo inicio) ---
    tl = load("course/.timeline_index.json")
    blocks = tl if isinstance(tl, list) else tl.get("blocks", [])
    bs = sorted(blocks, key=lambda b: b.get("period_start", ""))
    starts = [(b.get("id"), b.get("period_start")) for b in bs]

    def block_for_date(d_iso):
        cur = None
        for bid, st in starts:
            if st <= d_iso:
                cur = bid
            else:
                break
        return cur

    # --- TESTE-UNIDADE da borda (LOAD-BEARING). data==inicio -> esse bloco, nao anterior ---
    bar = "=" * 72
    print(bar); print("TESTE-UNIDADE BORDA [inicio, fim)"); print(bar)
    failures = []
    for bid, st in starts:
        got = block_for_date(st)
        if got != bid:
            failures.append((bid, st, got))
    if failures:
        print("FALHOU -- a borda nao e [inicio,fim). HALT.")
        for bid, st, got in failures:
            print(f"  data {st} (inicio de {bid}) caiu em {got}")
        sys.exit(1)
    print(f"OK: {len(starts)} blocos, toda data==inicio cai no proprio bloco (nao no anterior).")

    # --- uuid <-> display (pro predicted temporal) ---
    bi = load("course/.block_identity.json")
    u2d = {str(b.get("uuid")): str(b.get("display_id_last")) for b in bi}

    def disp(uuid_or_id):
        s = str(uuid_or_id or "").strip()
        return u2d.get(s, s)

    def predicted_temporal(e):
        t = str(e.get("temporal_block_id") or "").strip()
        if t:
            return disp(t)
        man = str(e.get("manual_timeline_block_id") or "").strip()
        if man:
            return disp(man)
        return disp(e.get("computed_block_id"))

    # --- manifest join index ---
    def nalnum(s):
        return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

    m = load("manifest.json")
    by_base, by_title, by_nbase, by_ntitle = {}, {}, {}, {}
    for e in m["entries"]:
        sp = str(e.get("source_path") or "")
        by_base[os.path.basename(sp).lower()] = e
        by_title[str(e.get("title") or "").lower()] = e
        by_nbase[nalnum(os.path.basename(sp))] = e
        by_ntitle[nalnum(e.get("title"))] = e

    def join(material):
        # EXATO (basename/title), depois NORMALIZADO-alnum (colapsa acento/mojibake/
        # pontuacao; EXATO na chave normalizada, NAO substring -> sem o force-match do
        # fuzzy). Sem match = NAO esta na manifest (podado/renomeado) -> unjoined honesto.
        k = str(material or "").strip().lower()
        nk = nalnum(material)
        return by_base.get(k) or by_title.get(k) or by_nbase.get(nk) or by_ntitle.get(nk)

    # --- Gabarito Subtopicos: subtopico# -> datas, fonte, umbrella ---
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    gab = list(wb["Gabarito Subtopicos"].iter_rows(values_only=True))[1:]
    sub_dates, sub_fonte = {}, {}
    for r in gab:
        if r[0] is None:
            continue
        n = int(r[0])
        dates = re.findall(r"\d{1,2}/\d{1,2}", str(r[2] or ""))
        sub_dates[n] = dates
        sub_fonte[n] = str(r[4] or "")

    # --- CROSSWALK: subtopico -> {blocos} via datas sob [inicio,fim) ---
    crosswalk = {}
    for n, dates in sorted(sub_dates.items()):
        bset = [block_for_date(iso(d)) for d in dates]
        crosswalk[n] = {"dates": dates, "blocks": bset,
                        "single": len(set(bset)) == 1, "fonte": sub_fonte[n]}

    # --- Rotulagem: por material ---
    rot = list(wb["Rotulagem"].iter_rows(values_only=True))[1:]
    out_rows, unjoined, straddle_flag = [], [], []
    for r in rot:
        if not any(c not in (None, "") for c in r):
            continue
        material = r[1]; bloco_correto = str(r[4] or ""); obs = str(r[6] or "")
        if material is None:
            continue
        # subtopico#: bloco_correto "NN ..." (existentes) ou obs "subtopico N" (notebooks)
        mnum = re.match(r"\s*(\d+)", bloco_correto)
        if mnum:
            sub = int(mnum.group(1))
        else:
            ms = re.search(r"subtopico\s+(\d+)", obs)
            sub = int(ms.group(1)) if ms else None
        scope = "clean" if (sub in CLEAN_SUBTOPICS) else ("eco-excluido" if sub in {12,13,14,15} else "sem-subtopico")
        fonte = sub_fonte.get(sub, "")

        # data_real: notebook (obs "-> DD/MM") direto; existente via subtopico
        md = re.search(r"->\s*(\d{1,2}/\d{1,2})", obs)
        prov = "?"
        if md:
            data_real = md.group(1); true_block = block_for_date(iso(data_real)); prov = f"obs:{fonte}"
        elif sub in crosswalk:
            cw = crosswalk[sub]
            if cw["single"]:
                data_real = cw["dates"][0] if cw["dates"] else ""; true_block = cw["blocks"][0]; prov = f"subt-single:{fonte}"
            else:
                data_real = ""; true_block = ""; prov = f"subt-STRADDLE:{fonte}"
        else:
            data_real = ""; true_block = ""

        e = join(material)
        eid = str(e.get("id")) if e else ""
        comp = disp(e.get("computed_block_id")) if e else ""
        temp = predicted_temporal(e) if e else ""
        pair = PAIRS.get(eid, "")
        if not e:
            unjoined.append(material)

        is_straddle = prov.startswith("subt-STRADDLE")
        scorable = bool(e) and scope == "clean" and bool(true_block) and not is_straddle
        discrim = scorable and (true_block != temp)
        if is_straddle and scope == "clean":
            straddle_flag.append((material, sub))

        out_rows.append({
            # true_block_id e o RÓTULO DE PONTUAÇÃO -> so para scorable. Non-scorable
            # (eco/straddle/unjoined) fica vazio pro eval pular (eval scoreia id+true).
            "id": eid, "material": material, "true_block_id": (true_block if scorable else ""),
            "computed_block_id": comp, "temporal_block_id": temp, "pair_key": pair,
            "provenance": prov, "scope": scope, "data_real": data_real,
            "scorable": "yes" if scorable else "no",
            "discriminante": "yes" if discrim else "no",
        })

    # --- CROSSWALK display ---
    print("\n" + bar); print("CROSSWALK  subtopico -> bloco(s)  (sob [inicio,fim))"); print(bar)
    for n in sorted(crosswalk):
        cw = crosswalk[n]
        tag = "single " if cw["single"] else "STRADDLE"
        sc = "clean" if n in CLEAN_SUBTOPICS else ("ECO" if n in {12,13,14,15} else "-")
        print(f"  subt {n:2} [{sc:5}] {tag} | datas {cw['dates']} -> {cw['blocks']} | fonte: {cw['fonte']}")

    # --- DENOMINADOR derivado ---
    scor = [x for x in out_rows if x["scorable"] == "yes"]
    disc = [x for x in scor if x["discriminante"] == "yes"]
    print("\n" + bar); print("DENOMINADOR DERIVADO (explicito)"); print(bar)
    print(f"  linhas Rotulagem            : {len(out_rows)}")
    print(f"  joined ao manifest          : {sum(1 for x in out_rows if x['id'])}/{len(out_rows)}  (unjoined: {len(unjoined)})")
    print(f"  escopo clean                : {sum(1 for x in out_rows if x['scope']=='clean')}")
    print(f"  eco-excluido (12-15)        : {sum(1 for x in out_rows if x['scope']=='eco-excluido')}")
    print(f"  straddle-flag (clean s/data): {len(straddle_flag)}  <- precisam data_real por-material [USER]")
    print(f"  >>> SCORABLE (denominador)  : {len(scor)}")
    print(f"  >>> DISCRIMINANTES          : {len(disc)}")

    print("\n  DISCRIMINANTES nomeados (true != temporal):")
    for x in disc:
        print(f"    FAIL {x['material'][:46]:46} data_real {x['data_real']} -> true {x['true_block_id']} | temporal {x['temporal_block_id']}")

    if straddle_flag:
        print("\n  STRADDLE-FLAG (clean, subtopico atravessa fronteira, sem data_real por-material):")
        for mat, sub in straddle_flag:
            print(f"    subt {sub} | {mat[:54]}")

    if unjoined:
        print("\n  UNJOINED (material sem match no manifest -- sem gap silencioso):")
        for u in unjoined:
            print(f"    {u}")

    # --- emite CSV (UNCOMMITTED) ---
    cols = ["id","material","true_block_id","computed_block_id","temporal_block_id",
            "pair_key","provenance","scope","data_real","scorable","discriminante"]
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for x in out_rows:
            w.writerow(x)

    print("\n" + bar)
    print(f"HALT pos-crosswalk. CSV escrito (NAO commitado): {OUT}")
    print("Revisa o crosswalk + denominador acima ANTES de rodar o eval. Nada commitado.")
    print(bar)


if __name__ == "__main__":
    main()
