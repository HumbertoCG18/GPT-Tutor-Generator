#!/usr/bin/env python3
"""Gerador de scaffold do gold material->bloco (planilha de rotulagem CEGA).

Lê o manifest + cronograma + block_identity de um repo-tutor (READ-ONLY) e
escreve um xlsx de rotulagem vazio para um humano preencher. Reverte o build-à-mão.

FIREWALL (não-negociável): a Aba 1 ("Rotulagem") NÃO contém, em lugar nenhum, o
bloco que o sistema computou para cada material — nada de pré-preencher, dica ou
coluna escondida. A coluna `tópico` deriva do NOME DO ARQUIVO (fallback: 1ª linha
de conteúdo), NUNCA da classificação de placement (concept_resolver/content_taxonomy).
Se isso for cruzado, o número do gold vale ~95% por construção e não serve.

Uso:
    python scripts/build_gold_xlsx.py --curso IA
    python scripts/build_gold_xlsx.py --curso IA --repo "C:/path/para/repo"

Saída: docs/reports/gold_templates/gold_<curso>_rotular.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- localização dos repos-tutor (default; sobrescrevível por --repo) ---------
GITHUB_BASE = Path.home() / "Documents" / "GitHub"
CURSO_REPOS = {
    "IA": "Inteligencia-Artifical-Tutor",
    "ES2": "Engenharia-Software-2-Tutor",
    "MF": "Metodos-Formais-Tutor",
    "SO": "Sistemas-Operacionais-Tutor",
    "TCC": "TCC-Tutor",
}

# --- clusters de tópico POR CURSO (opt-in) ------------------------------------
# Curso COM taxonomia própria agrupa a Aba 1 por tópico pedagógico (IA abaixo).
# Curso SEM entrada agrupa por `tipo` do material (genérico, firewall-safe) — não
# inventamos taxonomia por-curso que não dá pra validar (= chute confiante).
_IA_CLUSTER_ORDER = [
    "Intro / IA",
    "Dados",
    "Intro ML",
    "Supervisionado (kNN/redes/árvores)",
    "Avaliação",
    "Clustering",
    "Busca / AG / Minimax",
    "Provas / Listas",
    "Outros",
]
# Ordem de ATRIBUIÇÃO (específico vence genérico): cada material cai no 1º cluster
# cujos tokens casarem. Derivado do NOME do material — não do placement.
_IA_CLUSTER_RULES = [
    ("Avaliação", ("analisar-resultados", "analisarresultados", "avaliacao", "medidas-de-avaliacao",
                   "acc-pr", "acc, pr", " f1", "-f1", "sse")),
    ("Clustering", ("agrupamento", "cluster", "k-means", "kmeans", "hierarquico", "particional",
                    "nao-supervisionado", "naosupervisionado", "nao supervisionado")),
    ("Supervisionado (kNN/redes/árvores)", ("knn", "k-nn", "perceptron", "redesneurais", "redes-neurais",
                    "redes neurais", "rede-perceptron", "neural", "mlp", "arvore", "arvores", "decisao",
                    "classificacao", "supervisionado")),
    ("Busca / AG / Minimax", ("busca", "genetico", "geneticos", "minimax", "dijkstra", "hill-climbing",
                    "hc-sa", "hc, sa", "simulated", "annealing", "operadores", "programa-exemplo",
                    "programas-exemplo", "ag-feito", "algoritmo-genetico")),
    ("Intro ML", ("introducaoml", "introducao-a-ml", "introducao a ml", "intro-ml", "introml")),
    ("Dados", ("caracteristica", "exploratoria", "preparacao", "tipos-dados", "dados")),
    ("Provas / Listas", ("prova", "p1-", "p1_", "-p1", "lista")),
    ("Intro / IA", ("inteligencia", "historico", "visao-geral", "oracle", "responsavel",
                    "introducao-ia", "aula01", "-ia", " ia")),
]
# Registry curso -> (ordem de exibição, regras). Ausente = agrupa por `tipo`.
_TOPIC_CLUSTERS = {"IA": (_IA_CLUSTER_ORDER, _IA_CLUSTER_RULES)}
# Ordem dos grupos genéricos (por `tipo` de tipo_of). Tipo fora da lista vai ao
# fim, alfabético.
_TIPO_ORDER = ["prova", "lista", "código", "artigo", "pdf", "link", "material"]


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _norm(s: str) -> str:
    return _strip_accents(str(s or "")).lower()


def humanize(stem: str) -> str:
    """Descritor legível a partir do stem do arquivo. Sem classificação."""
    s = re.sub(r"\.[a-z0-9]{1,5}$", "", str(stem or ""))   # tira extensão
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)             # camelCase -> espaço
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ============================ leitura READ-ONLY ===============================
def load_manifest(repo: Path) -> dict:
    return json.loads((repo / "manifest.json").read_text(encoding="utf-8"))


def load_blocks(repo: Path):
    """Retorna (blocks, uuid->display). READ-ONLY."""
    tl_path = repo / "course" / ".timeline_index.json"
    data = json.loads(tl_path.read_text(encoding="utf-8"))
    blocks = data if isinstance(data, list) else data.get("blocks", [])
    led_path = repo / "course" / ".block_identity.json"
    u2d = {}
    if led_path.exists():
        led = json.loads(led_path.read_text(encoding="utf-8"))
        recs = led if isinstance(led, list) else led.get("blocks", [])
        for r in recs:
            u, d = str(r.get("uuid") or ""), str(r.get("display_id_last") or "")
            if u and d:
                u2d[u] = d
    return blocks, u2d


def block_display(block: dict, u2d: dict) -> str:
    return u2d.get(str(block.get("block_uuid") or ""), str(block.get("id") or ""))


def load_subtopics(path: Path) -> list:
    """Menu de subtópicos do SARC VIVO (csv auditável, fora do código).
    Colunas: subtopico,datas,guarda_chuva,letivo. READ-ONLY."""
    import csv
    rows = []
    with open(path, encoding="utf-8-sig") as f:  # tolera BOM (Excel-friendly)
        for r in csv.DictReader(f):
            sub = str(r.get("subtopico", "") or "").strip()
            if not sub:
                continue
            rows.append({
                "subtopico": sub,
                "datas": str(r.get("datas", "") or "").strip(),
                "guarda_chuva": str(r.get("guarda_chuva", "") or "").strip(),
                "letivo": str(r.get("letivo", "yes") or "yes").strip().lower() in ("yes", "sim", "true", "1"),
                "fonte_data": str(r.get("fonte_data", "") or "").strip(),
            })
    return rows


# ============================ Aba 1 (material) ================================
def material_filename(entry: dict) -> str:
    """Nome do arquivo cru. basename(source_path); url/sem-path -> title."""
    ft = str(entry.get("file_type", "") or "")
    if ft != "url":
        for fld in ("source_path", "raw_target"):
            v = str(entry.get(fld, "") or "").strip().rstrip(".")
            if v:
                base = os.path.basename(v.replace("\\", "/")).split("?")[0]
                if base:
                    return base
    return str(entry.get("title", "") or entry.get("id", "")).strip()


def first_content_line(repo: Path, entry: dict) -> str:
    """Fallback de tópico: 1ª linha de conteúdo do markdown curado."""
    for fld in ("curated_markdown", "approved_markdown", "base_markdown"):
        rel = str(entry.get(fld, "") or "").strip()
        if not rel:
            continue
        p = repo / rel
        if p.exists():
            for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
                t = line.strip().lstrip("#").strip()
                if t and not t.startswith("<!--") and not t.startswith("<"):
                    return t[:80]
    return ""


def topico_from_filename(repo: Path, entry: dict) -> str:
    """tópico = descritor do NOME DO ARQUIVO (fallback 1ª linha de conteúdo).
    FIREWALL: jamais usa computed_block/concept/placement."""
    stem = re.sub(r"\.[a-z0-9]{1,5}$", "", material_filename(entry))
    h = humanize(stem)
    if h:                       # nome do arquivo (mesmo curto, ex. "MLP") vence
        return h
    return first_content_line(repo, entry) or str(entry.get("id", ""))


def tipo_of(entry: dict) -> str:
    cat = str(entry.get("category", "") or "")
    ft = str(entry.get("file_type", "") or "")
    idt = _norm(entry.get("id", "") + " " + entry.get("title", ""))
    if cat == "provas" or re.search(r"\bp[12]\b", idt) or "prova" in idt:
        return "prova"
    if cat == "listas" or "lista" in idt:
        return "lista"
    if cat in ("references", "bibliografia") or "artigo" in idt or "survey" in idt:
        return "artigo"
    if cat == "codigo-professor" or ft in ("zip", "code"):
        return "código"
    if ft == "url":
        return "link"
    if ft == "pdf":
        return "pdf"
    return ft or "material"


def cluster_of(entry: dict, curso: str) -> str:
    """Grupo da Aba 1. Curso com taxonomia -> tópico; senão -> tipo do material."""
    topic = _TOPIC_CLUSTERS.get(curso)
    if not topic:
        return tipo_of(entry)
    name = _norm(entry.get("id", "") + " " + entry.get("title", ""))
    for cluster, toks in topic[1]:
        if any(tok in name for tok in toks):
            return cluster
    return "Outros"


def cluster_order_for(curso: str, entries: list) -> list:
    """Ordem de exibição dos grupos. Taxonomia do curso, ou tipos presentes."""
    topic = _TOPIC_CLUSTERS.get(curso)
    if topic:
        return topic[0]
    present = {tipo_of(e) for e in entries}
    return [t for t in _TIPO_ORDER if t in present] + sorted(present - set(_TIPO_ORDER))


# ============================ Aba 2 (cronograma) =============================
def _fmt_date(d: str) -> str:
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(d or ""))
    return f"{m.group(3)}/{m.group(2)}" if m else str(d or "")


def _clean_session_label(lbl: str) -> str:
    s = re.sub(r"\baula\b", "", str(lbl or ""), flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip(" -·")
    return s


def block_cronograma_topic(block: dict) -> str:
    """tópico da Aba 2 = VERDADE DO CRONOGRAMA (labels de sessão do SARC),
    NÃO o kind/topic derivado pelo sistema. Mostra a realidade do que foi ensinado."""
    parts = []
    for s in block.get("sessions", []) or []:
        lab = _clean_session_label(s.get("label", ""))
        if lab:
            parts.append(f"{_fmt_date(s.get('date',''))} {lab}")
    if parts:
        return " · ".join(parts)
    # sem sessões legíveis: cai pro topic_text cru do cronograma
    return _clean_session_label(block.get("topic_text", "")) or "—"


def block_tipo_from_sessions(block: dict) -> str:
    """tipo da Aba 2 derivado dos labels de sessão (cronograma), NÃO de block['kind']
    (que carrega o bug do bloco-06). Heurística por palavra-chave."""
    labels = [str(s.get("label", "") or "") for s in block.get("sessions", []) or []]
    t = _norm(" ".join(labels))
    has_suspensao = "suspens" in t or "feriado" in t or "recesso" in t
    teaching = any(k in t for k in (
        "k means", "k-means", "hierarquico", "perceptron", "arvore", "busca", "minimax",
        "dijkstra", "genetico", "agrupamento", "redes", "aprendizado", "classificacao",
        "introducao", "algoritmos", "conceitua", "dados", "agentes", "ml ", "knn"))
    if "plano de ensino" in t or "plano ensino" in t or "apresentacao da disciplina" in t:
        base = "apresentação"
    elif any(k in t for k in ("prova", "p1", "p2", "avaliacao", "substituicao")):
        base = "avaliação"
    elif "es day" in t or "evento" in t:
        base = "evento"
    elif "divulgacao" in t:
        base = "resultados"
    elif "duvida" in t or "atendimento" in t:
        base = "atendimento"
    elif any(k in t for k in ("apresentacao", "trabalho", "entrega", "t1", "t2")):
        base = "trabalho"
    elif "planejamento" in t:
        base = "planejamento"
    elif teaching:
        base = "aula"
    elif "aula" in t:
        base = "aula"
    else:
        base = "—"
    if has_suspensao and base == "aula":
        return "aula (c/ suspensão)"
    if has_suspensao and base == "—":
        return "feriado/suspensão"
    return base


# ============================ build do xlsx ==================================
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")   # amarelo claro = preencher
HDR_FONT = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)


def build_workbook(repo: Path, curso: str, subtopics: list | None = None):
    manifest = load_manifest(repo)
    blocks, u2d = load_blocks(repo)
    entries = [e for e in manifest.get("entries", []) if str(e.get("id", "")).strip()]

    # ---- Aba 2: gabarito dos blocos (cronograma) ----
    block_rows = []
    for b in blocks:
        disp = block_display(b, u2d)
        block_rows.append({
            "id": disp,
            "datas": f"{_fmt_date(b.get('period_start',''))}–{_fmt_date(b.get('period_end',''))}",
            "topico": block_cronograma_topic(b),
            "tipo": block_tipo_from_sessions(b),
            "sort": int(re.sub(r"\D", "", disp) or 0),
        })
    block_rows.sort(key=lambda r: r["sort"])

    # ---- Aba 1: materiais (cego), agrupado por cluster ----
    cluster_order = cluster_order_for(curso, entries)
    disp_index = {c: i for i, c in enumerate(cluster_order)}
    mat_rows = []
    for e in entries:
        topico = topico_from_filename(repo, e)
        cl = cluster_of(e, curso)
        mat_rows.append({
            "material": material_filename(e),
            "tipo": tipo_of(e),
            "topico": topico,
            "cluster": cl,
            "ckey": disp_index.get(cl, 99),
        })
    mat_rows.sort(key=lambda r: (r["ckey"], _norm(r["topico"])))

    # FIREWALL assert: nenhuma chave de placement vazou pra Aba 1
    leak_keys = {"computed_block_id", "temporal_block_id", "manual_timeline_block_id",
                 "computed_block_band", "computed_block_confidence", "block", "bloco"}
    for r in mat_rows:
        assert not (set(_norm(k) for k in r) & leak_keys), "FIREWALL: chave de placement na Aba 1"
        assert "bloco-" not in _norm(r["material"] + r["topico"] + r["tipo"]), \
            "FIREWALL: referência de bloco apareceu em material/tópico"

    wb = Workbook()

    # ===== Aba 2 primeiro (a Aba 1 referencia o dropdown dela) =====
    # Preferência: MENU DE SUBTÓPICOS do SARC vivo (csv). Senão, blocos do cache
    # (marca-se que está stale). O dropdown só lista entradas LETIVAS.
    ws2 = wb.active
    NONLETIVO_FILL = PatternFill("solid", fgColor="F2F2F2")
    if subtopics:
        sheet2_name = "Gabarito Subtopicos"
        ws2.title = sheet2_name
        ws2.append(["#", "subtópico", "datas", "guarda-chuva", "fonte", "_opcoes"])
        for c in range(1, 6):
            cell = ws2.cell(1, c); cell.font = HDR_FONT; cell.fill = HDR_FILL
        # numeração sequencial por subtópico (id estável do menu, 1..N)
        opts = []
        for i, s in enumerate(subtopics, start=1):
            s["num"] = i
            ws2.append([i, s["subtopico"], s["datas"], s["guarda_chuva"], s.get("fonte_data", "")])
            if s["letivo"]:
                opts.append(f"{i:02d} — {s['subtopico']}")
            else:
                ws2.cell(ws2.max_row, 4, f"{s['guarda_chuva']} ⊘ fora do dropdown")
                for c in range(1, 6):
                    cc = ws2.cell(ws2.max_row, c)
                    cc.fill = NONLETIVO_FILL
                    cc.font = Font(italic=True, color="999999")
        opts += ["N/A", "não sei"]
        # FIREWALL do menu: nenhuma entrada é guarda-chuva puro; nenhum não-letivo no dropdown
        UMBRELLAS = {"ml", "supervisionada", "supervisionado", "nao supervisionada",
                     "nao-supervisionada", "nao supervisionado", "clustering"}
        for s in subtopics:
            assert _norm(s["subtopico"]) not in UMBRELLAS, f"guarda-chuva como entrada: {s['subtopico']!r}"
        for s in subtopics:
            if not s["letivo"]:
                assert not any(s["subtopico"] in o for o in opts), \
                    f"FIREWALL: não-letivo no dropdown: {s['subtopico']!r}"
        ws2.cell(1, 6, "_opcoes")
        for i, o in enumerate(opts, start=2):
            ws2.cell(i, 6, o)
        ws2.column_dimensions["A"].width = 5
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 46
        ws2.column_dimensions["D"].width = 30
        ws2.column_dimensions["E"].width = 22
        ws2.column_dimensions["F"].hidden = True
        # wrap nas colunas de texto: nunca vaza pra célula vizinha
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.freeze_panes = "A2"
        menu_count = sum(1 for s in subtopics if s["letivo"])
        menu_total = len(subtopics)
    else:
        sheet2_name = "Gabarito dos Blocos"
        ws2.title = sheet2_name
        ws2.append(["bloco_id", "datas", "tópico (cache STALE)", "tipo", "", "_opcoes"])
        for c in range(1, 5):
            cell = ws2.cell(1, c); cell.font = HDR_FONT; cell.fill = HDR_FILL
        for r in block_rows:
            ws2.append([r["id"], r["datas"], r["topico"], r["tipo"]])
        opts = [r["id"] for r in block_rows] + ["N/A", "não sei"]
        ws2.cell(1, 6, "_opcoes")
        for i, o in enumerate(opts, start=2):
            ws2.cell(i, 6, o)
        ws2.column_dimensions["A"].width = 11
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 78
        ws2.column_dimensions["D"].width = 16
        ws2.column_dimensions["F"].hidden = True
        ws2.freeze_panes = "A2"
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=3, max_col=3):
            row[0].alignment = WRAP
        menu_count = len(block_rows)
        menu_total = len(block_rows)
    opts_last = 1 + len(opts)

    # ===== Aba 1: Rotulagem =====
    ws1 = wb.create_sheet("Rotulagem", 0)
    headers = ["#", "material", "tipo", "tópico", "bloco_correto", "confiança", "obs"]
    ws1.append(headers)
    for c, _ in enumerate(headers, start=1):
        cell = ws1.cell(1, c)
        cell.font = HDR_FONT
        cell.fill = EDIT_FILL if c in (5, 6, 7) else HDR_FILL
        cell.protection = LOCKED
    for i, r in enumerate(mat_rows, start=1):
        row = i + 1
        ws1.cell(row, 1, i).protection = LOCKED
        ws1.cell(row, 2, r["material"]).protection = LOCKED
        ws1.cell(row, 3, r["tipo"]).protection = LOCKED
        ws1.cell(row, 4, r["topico"]).protection = LOCKED
        # editáveis (vazias, destravadas)
        for c in (5, 6, 7):
            cell = ws1.cell(row, c, "")
            cell.protection = UNLOCKED
            cell.fill = EDIT_FILL
        # wrap nas colunas de texto longas (material, tópico, obs); resto top
        for c in range(1, 8):
            ws1.cell(row, c).alignment = WRAP if c in (2, 4, 7) else TOP
    last = len(mat_rows) + 1

    # largura da coluna do dropdown = maior opção ("NN — subtópico …")
    e_width = min(54, max((len(o) for o in opts), default=16) + 3)
    widths = {"A": 5, "B": 52, "C": 12, "D": 46, "E": e_width, "F": 12, "G": 34}
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # dropdown bloco_correto (col E) <- Aba 2 col F (blocos + N/A + não sei)
    dv_block = DataValidation(
        type="list",
        formula1=f"'{sheet2_name}'!$F$2:$F${opts_last}",
        allow_blank=True, showDropDown=False,
    )
    dv_block.error = "Escolha um subtópico da lista (ou N/A / não sei)."
    dv_block.prompt = "Selecione o subtópico correto (SARC)."
    ws1.add_data_validation(dv_block)
    dv_block.add(f"E2:E{last}")

    # dropdown confiança (col F)
    dv_conf = DataValidation(type="list", formula1='"alta,média,baixa"',
                             allow_blank=True, showDropDown=False)
    dv_conf.prompt = "Sua confiança na resposta."
    ws1.add_data_validation(dv_conf)
    dv_conf.add(f"F2:F{last}")

    # proteção: trava descritores, deixa E/F/G editáveis
    ws1.protection.sheet = True
    ws1.protection.selectLockedCells = True
    ws1.protection.selectUnlockedCells = True
    ws1.protection.formatColumns = False
    ws1.protection.sort = False
    ws1.protection.autoFilter = False

    return wb, mat_rows, menu_count, menu_total, cluster_order


def main(argv) -> int:
    ap = argparse.ArgumentParser(description="Gera scaffold CEGO do gold material->bloco.")
    ap.add_argument("--curso", required=True, help="IA / ES2 / MF / SO / TCC")
    ap.add_argument("--repo", default="", help="override do caminho do repo-tutor")
    ap.add_argument("--out", default="", help="override do caminho de saída do xlsx")
    ap.add_argument("--subtopics", default="", help="csv de subtópicos do SARC vivo "
                    "(default: docs/reports/gold_templates/sarc_subtopics_<curso>.csv)")
    args = ap.parse_args(argv)

    curso = args.curso.strip().upper()
    if args.repo:
        repo = Path(args.repo)
    else:
        if curso not in CURSO_REPOS:
            print(f"[erro] curso {curso!r} desconhecido; use --repo. Conhecidos: {list(CURSO_REPOS)}")
            return 2
        repo = GITHUB_BASE / CURSO_REPOS[curso]
    if not (repo / "manifest.json").exists():
        print(f"[erro] manifest.json não encontrado em {repo}")
        return 2

    templates_dir = Path(__file__).resolve().parents[1] / "docs" / "reports" / "gold_templates"
    out = Path(args.out) if args.out else (templates_dir / f"gold_{curso}_rotular.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    # Cursos cujos blocos NAO alinham com a semana do roteiro (straddle alto) ->
    # forcados a modo-bloco mesmo com csv presente. MF: 10/18 semanas straddle.
    FORCE_BLOCK = {"MF"}
    sub_path = Path(args.subtopics) if args.subtopics else (templates_dir / f"sarc_subtopics_{curso}.csv")
    subtopics = (load_subtopics(sub_path)
                 if (sub_path.exists() and curso not in FORCE_BLOCK) else None)

    wb, mat_rows, menu_count, menu_total, cluster_order = build_workbook(repo, curso, subtopics)
    wb.save(out)

    # ---- resumo + confirmação de firewall ----
    from collections import Counter
    clusters = Counter(r["cluster"] for r in mat_rows)
    fonte = f"SUBTÓPICOS SARC ({sub_path.name})" if subtopics else "BLOCOS (cache STALE — sem csv de subtópicos)"
    print(f"OK  curso={curso}  repo={repo.name}")
    print(f"    materiais (Aba 1): {len(mat_rows)}")
    print(f"    menu (Aba 2): fonte={fonte}")
    if subtopics:
        nonlet = [s["subtopico"] for s in subtopics if not s["letivo"]]
        print(f"      entradas no dropdown (letivas): {menu_count}  |  total no menu: {menu_total}")
        print(f"      não-letivos EXCLUÍDOS do dropdown ({len(nonlet)}): {nonlet}")
        print("      FIREWALL menu OK: nenhuma entrada é guarda-chuva (ML/Supervisionada/Clustering);"
              " nenhuma suspensão/feriado no dropdown.")
    print(f"    saída: {out}")
    print("    clusters Aba 1:", {k: clusters[k] for k in cluster_order if clusters.get(k)})
    print("    FIREWALL Aba 1 OK: [#, material, tipo, tópico, bloco_correto(VAZIO),"
          " confiança(VAZIO), obs(VAZIO)]; nenhum bloco/subtópico pré-preenchido;"
          " tópico derivado do nome do arquivo.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
