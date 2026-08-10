"""Camada de edicao xlsx pro gold de unidades (campanha 2, Task 7).

build:  gera UM workbook com 1 aba por curso a partir do INDICE EM DISCO de cada
        repo-tutor: TODOS os blocos (pedido do user 2026-08-08 — linha do tempo
        sem buracos), com os fora-da-regua (source_kind: prova/trabalho)
        acinzentados e marcados em notes. true_unit com dropdown dos slugs reais
        (taxonomia via course_probe). Datas exibidas = span das sessoes LETIVAS
        (borda suspensao/feriado fora — caso IA bloco-06). Se o workbook ja
        existe, PRESERVA true_unit/notes preenchidos (chave: block_uuid).
export: le o workbook editado e reescreve os CSVs canonico (utf-8-sig).
        Linhas fora-da-regua saem com true_unit vazio (eval ignora).

Uso:
    python scripts/gold_units_xlsx.py build
    python scripts/gold_units_xlsx.py export
"""
from __future__ import annotations

import csv
import json
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill, Protection  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from src.models.core import SubjectStore  # noqa: E402
from scripts.course_probe import compute_production_taxonomy  # noqa: E402
from scripts.eval_units import sigla_for_repo  # noqa: E402
from scripts.gold_units_template import FIELDS, OUT_DIR  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX = OUT_DIR / "gold_units_rotular.xlsx"
ORDER = ["MF", "SO", "ES2", "IA", "TCC"]
FORA_NOTE = "FORA DA RÉGUA"

# Correcoes de EXIBICAO de kind provadas na varredura 2026-08-08 (classifier de
# review por keyword nua — item [CODE] no tracker via T13). Morrem com o fix real.
DISPLAY_KIND_FIX = {
    ("TCC", "bloco-05"): "class",   # "Revisão: Chomsky..." = conteudo, dist 12 da prova
    ("TCC", "bloco-16"): "review",  # "revisão para prova p1", vespera real
    ("TCC", "bloco-26"): "review",  # "revisão para prova p2", vespera real
}

HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")   # amarelo = preencher
FORA_FILL = PatternFill("solid", fgColor="F2F2F2")   # cinza = fora da regua
HDR_FONT = Font(bold=True)
FORA_FONT = Font(italic=True, color="999999")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

COL_TRUE = FIELDS.index("true_unit") + 1
COL_NOTES = FIELDS.index("notes") + 1
WIDTHS = {"block_uuid": 4, "block_id": 10, "date_start": 11, "date_end": 11,
          "kind": 11, "topic_text": 58, "unit_slug_atual": 40, "true_unit": 46,
          "notes": 30}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).lower()


def _lective_span(block: dict) -> tuple[str, str]:
    """Span das sessoes letivas (suspensao/feriado/recesso de borda fora)."""
    dates = [str(s.get("date", ""))[:10] for s in (block.get("sessions") or [])
             if not any(k in _norm(s.get("label", "")) for k in ("suspensao", "feriado", "recesso"))]
    if not dates:  # bloco 100% nao-letivo (feriado puro): usa period original
        return str(block.get("period_start", ""))[:10], str(block.get("period_end", ""))[:10]
    return min(dates), max(dates)


def _courses():
    store = SubjectStore()
    out = []
    for name in store.names():
        sp = store.get(name)
        if sp is None or not getattr(sp, "repo_root", ""):
            continue
        sig = sigla_for_repo(Path(sp.repo_root))
        if sig:
            out.append((sig, sp))
    out.sort(key=lambda t: ORDER.index(t[0]) if t[0] in ORDER else 99)
    return out


def _harvest_labels() -> dict:
    """{(sheet, block_uuid): (true_unit, notes)} do workbook existente."""
    if not XLSX.exists():
        return {}
    wb = load_workbook(XLSX, data_only=True)
    out = {}
    for sig in ORDER:
        if sig not in wb.sheetnames:
            continue
        ws = wb[sig]
        hdr = [c.value for c in ws[1]]
        try:
            cu = hdr.index("block_uuid") + 1
            ct = hdr.index("true_unit") + 1
            cn = hdr.index("notes") + 1
        except ValueError:
            continue
        for r in range(2, ws.max_row + 1):
            uuid = str(ws.cell(r, cu).value or "").strip()
            tu = str(ws.cell(r, ct).value or "").strip()
            nt = str(ws.cell(r, cn).value or "").strip()
            if uuid and (tu or nt):
                out[(sig, uuid)] = (tu, nt)
    return out


def build() -> int:
    saved = _harvest_labels()
    wb = Workbook()
    wb.remove(wb.active)

    ws_s = wb.create_sheet("_slugs")
    ranges = {}
    courses = _courses()
    for col, (sig, sp) in enumerate(courses, start=1):
        tax = compute_production_taxonomy(sp)
        slugs = [u.get("slug") for u in tax.get("units", []) if u.get("slug")]
        ws_s.cell(1, col, sig)
        for i, slug in enumerate(slugs, start=2):
            ws_s.cell(i, col, slug)
        letter = get_column_letter(col)
        ranges[sig] = f"'_slugs'!${letter}$2:${letter}${1 + len(slugs)}"
    ws_s.sheet_state = "hidden"

    total = kept = 0
    for sig, sp in courses:
        idx = json.loads((Path(sp.repo_root) / "course" / ".timeline_index.json")
                         .read_text(encoding="utf-8"))
        blocks = idx.get("blocks", [])
        ws = wb.create_sheet(sig)
        ws.append(FIELDS)
        for c in range(1, len(FIELDS) + 1):
            cell = ws.cell(1, c)
            cell.font = HDR_FONT
            cell.fill = EDIT_FILL if c in (COL_TRUE, COL_NOTES) else HDR_FILL
        def _emit(vals, fora, editable_ok=True):
            ws.append(vals)
            row = ws.max_row
            for c in range(1, len(FIELDS) + 1):
                cell = ws.cell(row, c)
                editable = (c in (COL_TRUE, COL_NOTES)) and not fora and editable_ok
                cell.protection = UNLOCKED if editable else LOCKED
                if fora:
                    cell.fill = FORA_FILL
                    cell.font = FORA_FONT
                elif editable:
                    cell.fill = EDIT_FILL
                cell.alignment = WRAP if FIELDS[c - 1] in ("topic_text", "notes") else TOP

        NONLECT = ("suspensao", "feriado", "recesso")
        for b in blocks:
            fora = bool(b.get("source_kind"))
            ds, de = _lective_span(b)
            kind = DISPLAY_KIND_FIX.get((sig, b.get("id")), b.get("kind", ""))
            uuid = str(b.get("block_uuid") or "")
            tu, nt = saved.get((sig, uuid), ("", ""))
            if fora and not nt:
                nt = f"{FORA_NOTE} (kind={kind})"
            if tu:
                kept += 1
            # sessao NAO-letiva embutida em bloco de aula -> linha propria
            # sintetica (pedido user 2026-08-08: planilha espelha o CRONOGRAMA,
            # nao a segmentacao do indice; caso IA bloco-06). Sem uuid: nunca
            # exporta/rotula. topic_text do bloco fica sem os tokens nao-letivos.
            embedded = [s for s in (b.get("sessions") or [])
                        if any(k in _norm(s.get("label", "")) for k in NONLECT)]
            tt = str(b.get("topic_text", "") or "")
            if embedded and not fora and len(embedded) < len(b.get("sessions") or []):
                for s in sorted(embedded, key=lambda x: str(x.get("date", ""))):
                    d = str(s.get("date", ""))[:10]
                    lab = str(s.get("label", "") or "").strip()
                    _emit(["", f"({b.get('id', '')})", d, d, "suspended",
                           lab, "", "", f"{FORA_NOTE} (sessão não-letiva dentro de "
                           f"{b.get('id', '')} no índice; bloco próprio = fix de "
                           "segmentação pendente)"], fora=True)
                tt = " ".join(w for w in tt.split() if _norm(w) not in NONLECT)
            _emit([uuid, b.get("id", ""), ds, de, kind, tt,
                   str(b.get("unit_slug", "") or ""), tu, nt], fora=fora)
        for c, name in enumerate(FIELDS, start=1):
            ws.column_dimensions[get_column_letter(c)].width = WIDTHS[name]
        ws.column_dimensions["A"].hidden = True  # block_uuid: chave, nao mexer
        ws.freeze_panes = "A2"
        dv = DataValidation(type="list", formula1=ranges[sig],
                            allow_blank=True, showDropDown=False)
        dv.error = "Escolha um slug da lista (ou deixe vazio = fora da regua)."
        dv.prompt = "Slug da unidade CERTA deste bloco."
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(COL_TRUE)}2:{get_column_letter(COL_TRUE)}{ws.max_row}")
        ws.protection.sheet = True
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True
        n_gold = sum(1 for b in blocks if not b.get("source_kind"))
        total += n_gold
        print(f"  {sig}: {len(blocks)} blocos ({n_gold} na regua)")

    wb.save(XLSX)
    print(f"OK  {total} blocos-regua -> {XLSX} | rotulos preservados: {kept}")
    return 0


def export() -> int:
    wb = load_workbook(XLSX, data_only=True)
    for sig in ORDER:
        if sig not in wb.sheetnames:
            continue
        ws = wb[sig]
        headers = [c.value for c in ws[1]][: len(FIELDS)]
        if headers != FIELDS:
            print(f"[erro] {sig}: header do xlsx != FIELDS ({headers})")
            return 2
        out = OUT_DIR / f"gold_units_{sig}.csv"
        n = fill = 0
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(FIELDS)
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = ["" if v is None else str(v).strip() for v in row[: len(FIELDS)]]
                if not any(vals):
                    continue
                if not vals[0]:
                    continue  # linha sintetica (sessao nao-letiva) — sem uuid, nao exporta
                if vals[COL_NOTES - 1].startswith(FORA_NOTE):
                    vals[COL_TRUE - 1] = ""  # fora da regua nunca leva rotulo
                w.writerow(vals)
                n += 1
                if vals[COL_TRUE - 1]:
                    fill += 1
        print(f"  {sig}: {n} linhas exportadas, {fill} com true_unit -> {out.name}")
    return 0


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else ""
    if mode == "build":
        raise SystemExit(build())
    if mode == "export":
        raise SystemExit(export())
    print(__doc__)
    raise SystemExit(2)
